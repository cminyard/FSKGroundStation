#!/usr/bin/python3

# This program read a Kicad BOM CSV output from stdin or from a file
# supplied on the command line and generates a BOM in XLS format
# suitable for JLCPCB on standard output.  This involves rearranging
# the columns and changing the column names and renaming some
# footprints.

import sys
import csv
from openpyxl import Workbook

infn = None
outfn = None
do_stage2_xlat = False
use_murata = False

unknown_components = []

in_flags = True
for i in sys.argv[1:]:
    if in_flags and i.startswith('-'):
        if i == '--':
            in_flags = False
        elif i == '-2':
            do_stage2_xlat = True
        elif i == '-murata':
            use_murata = True
        else:
            sys.stderr.write("Unknown flag: " + i)
            sys.exit(1)
            pass
        pass
    else:
        in_flags = False
        if infn is None:
            infn = i
        elif outfn is None:
            outfn = i
        else:
            sys.stderr.write("Too many filenames given\n")
            sys.exit(1)
            pass
        pass
    pass

if outfn is None:
    sys.stderr.write("No CSV BOM file and XLS/CSV output file given\n")
    sys.exit(1)
    pass

if not infn.endswith(".csv"):
    sys.stderr.write("First file doesn't end in '.csv': " + infn[1] + "\n")
    sys.exit(1)
    pass

do_xls = True
if outfn.endswith(".csv"):
    do_xls = False
elif not outfn.endswith(".xls") and not outfn.endswith(".xlsx"):
    sys.stderr.write("Second file doesn't end in '.xls' or '.xlsx' or '.csv': "
                     + outfn + "\n")
    sys.exit(1)

f = open(infn)

cf = csv.reader(f, delimiter=';')
line = cf.__next__()
if len(line) != 7:
    sys.stderr.write("First line doesn't have 7 values, doesn't appear to be"
                     + " a Kicad BOM output")
    sys.exit(1)
    pass

expected_first_line = 'Id;Designator;Footprint;Quantity;Designation;Supplier and ref;'.split(";")
for i in range(0, len(expected_first_line)):
    if line[i] != expected_first_line[i]:
        sys.stderr.write("First line pos %d: Expected %s, got %s" %
                         (i, expected_first_line[i], line[i]))
        sys.exit(1)
        pass
    pass

footprint_xlats = {
    'R_0402_1005Metric': '0402',
    'R_0603_1608Metric': '0603',
    'R_0805_2012Metric': '0805',
    'R_1206_3216Metric': '1206',
    'C_0402_1005Metric': '0402',
    'C_0603_1608Metric': '0603',
    'C_0805_2012Metric': '0805',
    'C_1206_3216Metric': '1206',
    'L_0402_1005Metric': '0402',
    'L_0603_1608Metric': '0603',
    'L_0805_2012Metric': '0805',
    'L_1206_3216Metric': '1206',
    'D_0603_1608Metric': '0603',
    'D_SOD_323':         'SOD-323',
    'D_SOD_882':         'SOD-882',

    'C_0603_1608Metric_Pad1.08x0.95mm_HandSolder': '0603',
    'C_0805_2012Metric_Pad1.18x1.45mm_HandSolder': '0805',
    'R_0603_1608Metric_Pad0.98x0.95mm_HandSolder': '0603',
    'R_1206_3216Metric_Pad1.30x1.75mm_HandSolder': '1206',
    'L_0603_1608Metric_Pad1.05x0.95mm_HandSolder': '0603',
    'C_0402_1005Metric_Pad0.74x0.62mm_HandSolder': '0402',
    'LED_0603_1608Metric_Pad1.05x0.95mm_HandSolder': '0603',
    'C_1206_3216Metric_Pad1.33x1.80mm_HandSolder': '1206',
    'L_1210_3225Metric_Pad1.42x2.65mm_HandSolder': '1210',
    'L_0402_1005Metric_Pad0.77x0.64mm_HandSolder': '0402',
    'R_0402_1005Metric_Pad0.72x0.64mm_HandSolder': '0402',
    'L_0805_2012Metric_Pad1.05x1.20mm_HandSolder': '0805',
    'L_1008_2520Metric_Pad1.43x2.20mm_HandSolder': '1008',

    'SOT-23-5_HandSoldering': 'SOT-23-5',
    'SOT-23-6_HandSoldering': 'SOT-23-6',
    ('',	''): ('',	''),
}
def xlat_footprint(s):
    if s in footprint_xlats:
        return footprint_xlats[s]
    return s

# Mostly Coilcraft parts.  All are automotive certified and have
# somewhat higher Q values than the Murata parts.
value_to_partnum_xlats_1 = {
    ('',	''): ('',	''),
}

# These are all Murata parts for things not already Murata.
value_to_partnum_xlats_1b = {
    ('',	''): ('',	''),
}

# General passive parts.
value_to_partnum_xlats_2 = {
    ('2.2nH 2%',	'0402'): ('LQW15AN2N2G80D', 'Murata'),
    ('9.1nH 2%',	'0402'): ('LQW15AN9N1G00D', 'Murata'),
    ('10nH 2%',		'0402'): ('LQW15AN10NG00D', 'Murata'),
    ('15nH 2%',		'0402'): ('LQW15AN15NG00D', 'Murata'),
    ('20nH',		'0402'): ('LQW15AN20NG00D', 'Murata'),
    ('22nH 2%',		'0402'): ('LQG15HN22NG02D', 'Murata'),
    ('23nH',		'IND_0908SQ-23NJLC'): ('0908SQ-23NJLC',  'Coilcraft'),
    ('27nH 2%',		'0402'): ('LQW15AN27NG00D', 'Murata'),
    ('27nH 2%',		'0603'): ('LQW18AN27NG00D', 'Murata'),
    ('43nH 2%',		'0402'): ('LQW15AN43NG00D', 'Murata'),
    ('47nH 2%',		'0603'): ('LQW18AN47NG00D', 'Murata'),
    ('56nH',		'0603'): ('LQW18AN56NG00D', 'Murata'),
    ('68nH 2%',		'0603'): ('LQW18AN68NG00D', 'Murata'),
    ('82nH 2%',		'0603'): ('LQW18AN82NG00D', 'Murata'),
    ('150nH',		'0805'): ('LQW2BANR15G00L', 'Murata'),
    ('150nH 2%',	'0805'): ('LQW2BANR15G00L', 'Murata'),
    ('220nH 2%',	'0805'): ('LQW2BASR22G00L', 'Murata'),
    ('470nH',		'0805'): ('LQW21FTR47M0HL', 'Murata'),
    ('1uH',		'1008'): ('IMSC1008AZER1R0M', 'Vishay Dale'),
    ('6.8uH',		'1210'): ('LQH32PH6R8NNCL', 'Murata'),
    ('6.8uH 3A',	'IND_1255AY-6R8M=P3'): ('1264EY-6R8M=P3', 'Murata'),

    ('0Ω',		'0402'): ('ERJ2GE0R00X', 'Panasonic'),
    ('18Ω',		'0402'): ('ERJ2RKF18R0X', 'Panasonic'),
    ('22Ω',             '0402'): ('ERJ2RKF22R0X', 'Panasonic'),
    ('348Ω 1%',		'0402'): ('ERJ2RKF3480X', 'Panasonic'),
    ('1.96K 1%',	'0402'): ('ERJ2RKF1961X', 'Panasonic'),
    ('5.1K',		'0402'): ('ERJ2RKF5101X', 'Panasonic'),
    ('6.04K 1%',	'0402'): ('ERJ2RKF6041X', 'Panasonic'),
    ('10K',		'0402'): ('ERJ2RKF1002X', 'Panasonic'),
    ('10K 1%',		'0402'): ('ERJ2RKF1002X', 'Panasonic'),
    ('30K',		'0402'): ('ERJ2RKF3002X', 'Panasonic'),
    ('56K 1%',		'0402'): ('ERJPA2F5602X', 'Panasonic'),
    ('62K',		'0402'): ('ERJ2RKF6202X', 'Panasonic'),

    ('2pF 1%',		'0402'): ('GCM1555C1H2R0BA16D', 'Murata'),
    ('2.2pF 1%',	'0402'): ('GJM1555C1H2R2CB01D', 'Murata'),
    ('5pF ±.1',		'0402'): ('GCM1555C1H5R0FA16D', 'Murata'),
    ('5.1pF 1%',	'0402'): ('GJM1555C1H5R1BB01D', 'Murata'),
    ('5.6pF 1%',	'0402'): ('GJM1555C1H5R6BB01D', 'Murata'),
    ('6.2pF 1%',	'0402'): ('GJM1555C1H6R2BB01D', 'Murata'),
    ('10pF 1%',		'0402'): ('GCM1555C1H100FA16D', 'Murata'),
    ('12pF 1%',		'0402'): ('GJM1555C1H120FB01D', 'Murata'),
    ('14pF 1%',		'0402'): ('GJM1555C1H140FB01D', 'Murata'),
    ('15pF 1%',		'0402'): ('GRM1555C1H150FA01D', 'Murata'),
    ('18pF 1%',         '0402'): ('GRM1555C1H180FA01D', 'Murata'),
    ('27pF 1%',		'0402'): ('GJM1555C1H270FB01D', 'Murata'),
    ('33pF 1%',		'0402'): ('GRT1555C1H330FA02D', 'Murata'),
    ('39pF 1%',		'0402'): ('GCM1555C1H390FA16D', 'Murata'),
    ('47pF 1%',         '0402'): ('GJM1555C1H470FB01D', 'Murata'),
    ('56pF 1%',		'0402'): ('GRM1555C1H560FA01D', 'Murata'),
    ('100pF 1%',	'0402'): ('GCM1555C1H101FA16D', 'Murata'),
    ('100pF',		'0402'): ('GCM1555C1H101FA16D', 'Murata'),
    ('220pF',           '0402'): ('GRM1555C1H221JA01D', 'Murata'),
    ('270pF 1%',	'0402'): ('GCM1555C1H271FA16D', 'Murata'),
    ('680pF',		'0402'): ('GCM1555C1H681JA16D', 'Murata'),
    ('1nF',		'0402'): ('GCM1555C1H102JA16D', 'Murata'),
    ('10nF',		'0402'): ('GCM155R71E103KA37D', 'Murata'),
    ('47nF',		'0402'): ('GCM155R71H473KE02D', 'Murata'),
    ('100nF',		'0402'): ('GCM155R71C104KA55D', 'Murata'),
    ('220nF',		'0402'): ('CL05B224KO5NNNC', 'Samsung'),
    ('1uF',		'0603'): ('CL10A105KB8NNNC', 'Samsung'),
    ('10uF 10V',	'1206'): ('CL31A106KBHNNNE', 'Samsung'),
    ('10uF',		'1206'): ('CL31A106KBHNNNE', 'Samsung'),
    ('22uF',		'1206'): ('CL31A226KAHNNNE', 'Samsung'),
    ('',	''): ('',	''),
}

other_components = {
    ('PJ-063BH', 'PJ-063BH_CUD'): None,
    ('NTCG103JF103FTDS 10KΩ@25C', '0402'): None,
    ('SKYA21003',               'QFN12_SKYA21003_SKY'): None,
    ('CY7C65215-32LTXI',        'PG-VQFN-32-803'): None,
    ('TPSM861253RDXR',          'RDX0007A-MFG'): None,
    ('HTSW-102-07-G-S',         'CON2_1X2_TU_TSW'): None,
    ('GRF5604',                 'QFN-16-1EP_3x3mm_P0.5mm_EP1.7x1.7mm'): None,
    ('TPD4E004DRYR',            'DRY6'): None,
    ('74LVC1GU04DRL',           'SOT-553'): None,
    ('CC1125RHBR',              'RHB32_3P45X3P45_TEX'): None,
    ('HTSW-103-07-G-S',         'CON3_1X3_TU_TSW'): None,
    ('CX90B1-24P',              'CX90B1-24P_HIR'): None,
    ('CC1312R1F3RGZR',          'RGZ0048A'): None,
    ('FTSH-105-01-L-DV-K',      'FTSH-105-01-L-DV-K'): None,
    ('OW7EL89CENUXK7YLC-40M',   'Oscillator_SMD_SiT_PQFN-4Pin_2.0x1.6mm'): None,
    ('OW7EL89CENUXK7YLC-48M',   'Oscillator_SMD_SiT_PQFN-4Pin_2.0x1.6mm'): None,
    ('BWSMA-KWE-Z001',          'CONN_BWSMA-KWE-Z001_BAT'): None,
    ('BSS138', 'SOT-23'): None,
    ('SJ-3524-SMT', 'CUI_SJ-3524-SMT'): None,
    ('', ''): None,
}

def xlat_value_to_partnum(s, footprint):
    v = (s, footprint)
    if use_murata and v in value_to_partnum_xlats_1b:
        return value_to_partnum_xlats_1b[v]
    if v in value_to_partnum_xlats_1:
        return value_to_partnum_xlats_1[v]
    if do_stage2_xlat and v in value_to_partnum_xlats_2:
        return value_to_partnum_xlats_2[v]
    if v not in value_to_partnum_xlats_2 and v not in other_components:
        unknown_components.append(s + "; " + footprint)
    return ('', s)

if do_xls:
    # Output in Excel format
    wb = Workbook()
    ws = wb.active

    lineno = 1
    ws.cell(lineno, 1, 'Comment')
    ws.cell(lineno, 2, 'Designator')
    ws.cell(lineno, 3, 'Footprint')
    ws.cell(lineno, 4, 'Value')
    ws.cell(lineno, 5, 'Manufacturer')

    for line in cf:
        lineno += 1
        if len(line) != 8:
            sys.stderr.write("Line %s doesn't have 8 values, it has %d" %
                             (lineno, len(line)));
            sys.exit(1)
            pass
        designator = line[1]
        footprint = xlat_footprint(line[2]).strip('"')
        (partnum, mfg) = xlat_value_to_partnum(line[4], footprint)
        if partnum == "":
            partnum = line[4]
        else:
            partnum = partnum + " " + line[4]
            pass
        partnum = partnum.replace(' ', ',')
        ws.cell(lineno, 1, partnum)
        ws.cell(lineno, 2, designator)
        ws.cell(lineno, 3, footprint)
        ws.cell(lineno, 4, line[4])
        pass

    wb.save(outfn)
    pass
else:
    # Output in CSV format
    outfile = open(outfn, "w")
    ocf = csv.writer(outfile)
    lineno = 1
    ocf.writerow(('Comment', 'Designator', 'Footprint', 'Value',
                  'Manufacturer'))
    for line in cf:
        lineno += 1
        if len(line) != 8:
            sys.stderr.write("Line %s doesn't have 8 values, it has %d" %
                             (lineno, len(line)));
            sys.exit(1)
            pass
        designator = line[1]
        footprint = xlat_footprint(line[2]).strip('"')
        (partnum, mfg) = xlat_value_to_partnum(line[4], footprint)
        if partnum == "":
            partnum = line[4]
        else:
            partnum = partnum + " " + line[4]
            pass
        partnum = partnum.replace(' ', ',')
        #comment = comment.replace('Ω', 'ohm')
        ocf.writerow((partnum, designator, footprint, line[4], mfg))
        pass
    pass

if unknown_components:
    print("Unknown components:")
    for i in unknown_components:
        print("  " + i)
        pass
    pass
